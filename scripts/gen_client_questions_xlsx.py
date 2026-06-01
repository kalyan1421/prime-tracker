#!/usr/bin/env python3
"""Generate the Prime Tracker client-questions workbook (xlsx)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "docs/Prime-Tracker-Client-Questions.xlsx"

NAVY = "1F3A5F"
LIGHT = "EAF0F6"
P1, P2, P3 = "C0392B", "B9770E", "1E8449"

sections = {
    "1. Excel / Data Migration": [
        ("1.1", "Is there one master Excel file or several (per project, financials, sales)?", "Determines import scope and whether sources must be merged.", "P1"),
        ("1.2", "Can you share the file with column headers and 3-5 sample rows per tab?", "We map every column to a schema field.", "P1"),
        ("1.3", "What does each tab/sheet represent? Please label them.", "Each tab likely maps to one entity.", "P1"),
        ("1.4", "How are Projects > Buildings > Units represented? Is there a building column?", "Every Unit must belong to a Building in our model.", "P1"),
        ("1.5", "What is the unique identifier per project / unit (code, address, name)?", "Needed to de-duplicate and link rows across tabs.", "P1"),
        ("1.6", "What currency, number and date formats are used (e.g. $, MM/DD/YYYY)?", "Prevents misparsed amounts/dates on import.", "P1"),
        ("1.7", "Are money values gross or net? Any formula cells vs. raw values?", "We need the underlying numbers, not formulas.", "P2"),
        ("1.8", "How should blank / TBD / N/A cells be handled?", "Decide null vs. zero vs. skip.", "P2"),
        ("1.9", "One-time seed import, or ongoing updates expecting re-import/sync?", "Drives whether we build an upload+reconcile feature.", "P1"),
        ("1.10", "Who owns the spreadsheet and can answer column-level questions?", "Single source of truth for mapping.", "P2"),
        ("1.11", "Do you want an 'export to Excel' of any screen/report later?", "Build now vs. defer.", "P3"),
    ],
    "2. Projects & Structure": [
        ("2.1", "How many active projects today, and total incl. completed/sold?", "Sizing, pagination, filters.", "P2"),
        ("2.2", "Do statuses map to ACTIVE/ON_HOLD/COMPLETED/CANCELLED or other words?", "Align enums to your vocabulary.", "P1"),
        ("2.3", "Confirm phases (Pre-Dev..Sold-Refi) and types (Residential..Industrial).", "Confirm or extend enums.", "P1"),
        ("2.4", "Are unit types complete (Retail, Medical, Flex, Lot, Office, Restaurant, Event)?", "Avoid missing categories at import.", "P2"),
        ("2.5", "What does 'Prime owned' mean and how is it marked in the sheet?", "We have a primeOwned flag needing a rule.", "P2"),
    ],
    "3. Financials": [
        ("3.1", "Track at budget-line level by category? Are our 10 categories correct?", "Confirm budget taxonomy.", "P1"),
        ("3.2", "Where do actual spend figures come from - manual, Excel, or QuickBooks?", "Import vs. integration.", "P1"),
        ("3.3", "Do you use vendor commitments / POs? How tracked today?", "Confirm Commitments module fits.", "P2"),
        ("3.4", "What variance threshold flags a budget as over (default >10%)?", "Drives alerts.", "P2"),
    ],
    "4. Loans & Debt": [
        ("4.1", "Which loan types (Construction, Permanent, Bridge, Mezzanine, SBA)? Others?", "Confirm enum.", "P2"),
        ("4.2", "Which loan fields are sensitive and must be encrypted/restricted?", "We encrypt sensitive loan fields.", "P1"),
        ("4.3", "Do you track draw requests against construction loans?", "Confirm draw feature.", "P2"),
        ("4.4", "How far ahead to alert on loan maturity (default 60 days)?", "Notification timing.", "P3"),
    ],
    "5. Sales & Leads": [
        ("5.1", "Do sale stages match Prospect>LOI>Under Contract>Closed>Cancelled?", "Confirm pipeline.", "P1"),
        ("5.2", "Track Leads separately from Sales with a 'convert to sale' step?", "Confirm workflow.", "P2"),
        ("5.3", "What lead sources / activity types do you log?", "Configure activity timeline.", "P3"),
    ],
    "6. Leases & Rent Roll": [
        ("6.1", "Do lease statuses match Draft/Active/Expired/Terminated?", "Confirm enum.", "P2"),
        ("6.2", "How is rent structured - flat, $/sqft, escalations, CAM?", "Rent-roll accuracy.", "P1"),
        ("6.3", "How early to warn on lease expiry (we use 30 and 7 days)?", "Notification timing.", "P3"),
    ],
    "7. Users & Access": [
        ("7.1", "Confirm roles: Founder, Finance, PM, Sales, Construction, Viewer. Others?", "Maps to RBAC.", "P1"),
        ("7.2", "Everyone signs in via Google Workspace @primedevelopers.com? External users?", "OAuth domain restriction.", "P1"),
        ("7.3", "Who should have admin (Founder) rights to manage users?", "Initial admin setup.", "P2"),
        ("7.4", "Should Finance/Founder be required to use MFA?", "Enforce per role.", "P2"),
        ("7.5", "Which roles see loan/financial detail vs. just sales/units?", "Field-level permissions.", "P1"),
    ],
    "8. Reports & Notifications": [
        ("8.1", "Which reports matter most (portfolio, sales, revenue, debt)? Missing any?", "Prioritize report tabs.", "P2"),
        ("8.2", "What KPIs does leadership review weekly/monthly?", "Dashboard + KPI snapshots.", "P2"),
        ("8.3", "Email notifications, in-app only, or both?", "Affects SMTP setup.", "P2"),
        ("8.4", "What time/timezone for the daily digest (default 8 AM CT)?", "Scheduled job config.", "P3"),
    ],
    "9. Integrations": [
        ("9.1", "Use QuickBooks? Want live sync? Can you provide credentials?", "QB integration stubbed pending creds.", "P2"),
        ("9.2", "Any other systems to connect (banking, CRM, DocuSign, Drive)?", "Scope integrations.", "P3"),
        ("9.3", "Should documents be attachable to units/comments? Stored where today?", "File attachments not built yet.", "P3"),
    ],
    "10. Rollout": [
        ("10.1", "Target go-live date and any hard deadlines?", "Sequencing.", "P1"),
        ("10.2", "Who are first pilot users and the day-one must-have features?", "MVP scope.", "P1"),
        ("10.3", "Want a training/walkthrough session and quick-start guide?", "Adoption.", "P3"),
        ("10.4", "Any multi-tenant / multi-entity need soon (separate LLCs)?", "Multi-tenant hooks ready.", "P2"),
    ],
}

wb = Workbook()
ws = wb.active
ws.title = "Client Questions"

thin = Side(style="thin", color="D0D7DE")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
prio_color = {"P1": P1, "P2": P2, "P3": P3}

# Title
ws.merge_cells("A1:F1")
t = ws["A1"]
t.value = "Prime Tracker - Client Questions (Prime Developers)"
t.font = Font(bold=True, size=15, color="FFFFFF")
t.fill = PatternFill("solid", fgColor=NAVY)
t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:F2")
leg = ws["A2"]
leg.value = "Priority:  P1 = blocks build/import   |   P2 = needed soon   |   P3 = nice to clarify       Answer column is for the client."
leg.font = Font(italic=True, size=10, color="444444")
ws.row_dimensions[2].height = 18

headers = ["#", "Section", "Question", "Why we're asking", "Priority", "Client answer"]
hr = 3
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=hr, column=c, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
ws.row_dimensions[hr].height = 22

r = hr + 1
for section, rows in sections.items():
    for i, (num, q, why, prio) in enumerate(rows):
        ws.cell(row=r, column=1, value=num)
        ws.cell(row=r, column=2, value=section if i == 0 else "")
        ws.cell(row=r, column=3, value=q)
        ws.cell(row=r, column=4, value=why)
        pc = ws.cell(row=r, column=5, value=prio)
        pc.font = Font(bold=True, color=prio_color[prio])
        pc.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=6, value="")
        shade = LIGHT if (r % 2 == 0) else "FFFFFF"
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if c != 5:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            if c in (2,):
                cell.font = Font(bold=True, color=NAVY)
            if cell.fill.fgColor.rgb in (None, "00000000"):
                cell.fill = PatternFill("solid", fgColor=shade)
        r += 1

widths = {"A": 6, "B": 22, "C": 52, "D": 40, "E": 9, "F": 34}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

ws.freeze_panes = "A4"
ws.sheet_view.showGridLines = False
ws.print_options.horizontalCentered = True
ws.page_setup.fitToWidth = 1

wb.save(OUT)
print(f"Wrote {OUT} with {r - hr - 1} questions across {len(sections)} sections.")
